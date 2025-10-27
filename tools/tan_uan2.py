# 設定路徑（Windows 用戶需求）
# INPUT_PATH = r"c:\work\DoMiSo_Zu_Im\tools\python_in_excel.xlsx"
# OUTPUT_PATH = r"c:\work\DoMiSo_Zu_Im\tools\python_in_excel_with_conversion.xlsx"
# Python 套件安裝： pip install pandas openpyxl

# -*- coding: utf-8 -*-
import re

import pandas as pd
from caas_jupyter_tools import display_dataframe_to_user  # 只在沙盒顯示用
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Linux 版路徑設定
# INPUT_PATH = "/mnt/data/python_in_excel.xlsx"
# OUTPUT_PATH = "/mnt/data/python_in_excel_with_conversion.xlsx"
# Windows 版路徑設定
INPUT_PATH = r"c:\work\DoMiSo_Zu_Im\tools\python_in_excel.xlsx"
OUTPUT_PATH = r"c:\work\DoMiSo_Zu_Im\tools\python_in_excel_with_conversion.xlsx"
SHEET_NAME = "中州韻輸入方案字庫"

# 讀取：第一列當表頭
df = pd.read_excel(INPUT_PATH, sheet_name=SHEET_NAME, header=0)

# **立刻**忽略多餘的 Unnamed 欄位
df = df.loc[:, ~df.columns.str.contains(r"^Unnamed")]

# 確保有到 E 欄（避免插入位置越界；必要時補空欄）
while df.shape[1] < 5:
    df[f"__補欄{df.shape[1] + 1}__"] = None

# 在 D 與 E 之間插入新欄（0-based：A=0,B=1,C=2,D=3,E=4；在 index=4 插入）
INSERT_POS = 4
NEW_COL = "調整後台羅音標"
if NEW_COL not in df.columns:
    df.insert(INSERT_POS, NEW_COL, None)

# 取得 D 欄（index=3）
col_D = df.iloc[:, 3].astype(str)


# 規則：([ptkh])1 -> \1 0
def convert(text: str):
    new_text = re.sub(r"([ptkh])1", r"\g<1>0", text)
    return new_text, (new_text != text)


converted = col_D.apply(convert)
df[NEW_COL] = [c[0] for c in converted]
changed_flags = [c[1] for c in converted]

# 先輸出（不含格式）
df.to_excel(OUTPUT_PATH, sheet_name=SHEET_NAME, index=False)

# 用 openpyxl 加上樣式（只標有改動的格子）
wb = load_workbook(OUTPUT_PATH)
ws = wb[SHEET_NAME]

# 找新欄位實際位置（1-based）
new_col_idx = None
for i, cell in enumerate(ws[1], start=1):
    if cell.value == NEW_COL:
        new_col_idx = i
        break

fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
font = Font(color="FF0000")

for r, changed in enumerate(changed_flags, start=2):  # 跳過表頭
    if changed:
        cell = ws.cell(row=r, column=new_col_idx)
        cell.fill = fill
        cell.font = font

wb.save(OUTPUT_PATH)

# 沙盒顯示前 30 筆
display_cols = list(df.columns[:5])
display_dataframe_to_user("調整後台羅音標：預覽前 30 筆", df.loc[:29, display_cols])

{
    "input_path": INPUT_PATH,
    "output_path": OUTPUT_PATH,
    "sheet_name": SHEET_NAME,
    "total_rows": int(len(df)),
    "changed_count": int(sum(changed_flags)),
}
