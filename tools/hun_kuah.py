# -*- coding: utf-8 -*-
"""
hun_kuah.py - 將單一資料分成兩筆
將含有【/】符號的儲存格內容分割成多筆資料記錄

需求：
1. 檢查【調整後台羅音標】欄（E欄）是否含有【/】符號
2. 如有，則依【/】符號分割資料
3. 在原記錄後新增一筆資料
4. 調整序號並填入分割後的資料

Python 套件安裝： pip install pandas openpyxl
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# === 路徑自行調整 ===
INPUT_PATH = r"c:\work\DoMiSo_Zu_Im\tools\python_in_excel_with_conversion.xlsx"
OUTPUT_PATH = r"c:\work\DoMiSo_Zu_Im\tools\python_in_excel_split.xlsx"
SHEET_NAME = "中州韻輸入方案字庫"
TARGET_COL = "調整後台羅音標"  # E 欄

def process_split_data():
    """處理分割資料的主要函數"""

    # 讀檔：第一列為表頭
    df = pd.read_excel(INPUT_PATH, sheet_name=SHEET_NAME, header=0)

    # 立刻忽略多餘的 Unnamed 欄位
    df = df.loc[:, ~df.columns.str.contains(r"^Unnamed")]

    print(f"原始資料筆數：{len(df)}")
    print(f"欄位名稱：{list(df.columns)}")

    # 檢查目標欄位是否存在
    if TARGET_COL not in df.columns:
        print(f"錯誤：找不到目標欄位 '{TARGET_COL}'")
        return

    # 重設索引以便安全操作
    df = df.reset_index(drop=True)

    # 用來收集新增的資料列
    new_rows = []
    split_count = 0

    # 由後往前處理，避免索引位移問題
    for idx in reversed(range(len(df))):
        row = df.iloc[idx]
        target_value = str(row[TARGET_COL]) if pd.notna(row[TARGET_COL]) else ""

        # 檢查是否含有【/】符號
        if "/" in target_value:
            print(f"處理第 {idx+2} 列 (序號: {row.iloc[0]}): {target_value}")

            # 依【/】符號分割
            parts = [part.strip() for part in target_value.split("/")]

            if len(parts) >= 2:
                split_count += 1

                # 修改原記錄：使用第一個分割值
                df.loc[idx, TARGET_COL] = parts[0]

                # 為每个額外的分割值創建新記錄
                for i, part in enumerate(parts[1:], 1):
                    new_row = row.copy()

                    # 調整序號：原值改為【原值-分割序號】
                    if pd.notna(new_row.iloc[0]):  # A欄：序號
                        try:
                            original_seq = int(new_row.iloc[0])
                            new_row.iloc[0] = f"{original_seq}-{i}"
                        except (ValueError, TypeError):
                            new_row.iloc[0] = f"{new_row.iloc[0]}-{i}"

                    # 調整後台羅音標（E欄）：使用分割值
                    new_row[TARGET_COL] = part

                    # 備註（G欄）：清空
                    if len(new_row) > 6:  # 確保有G欄
                        new_row.iloc[6] = ""

                    # 記錄新增的資料列和插入位置
                    new_rows.append((idx + 1, new_row))

                print(f"  分割為 {len(parts)} 部分: {parts}")

    # 按索引順序插入新列（由大到小，避免索引位移）
    for insert_pos, new_row in sorted(new_rows, reverse=True):
        # 將新列插入到指定位置
        df_before = df.iloc[:insert_pos]
        df_after = df.iloc[insert_pos:]
        new_row_df = pd.DataFrame([new_row], columns=df.columns)
        df = pd.concat([df_before, new_row_df, df_after], ignore_index=True)

    print("處理完成：")
    print(f"- 分割筆數：{split_count}")
    print(f"- 新增資料列：{len(new_rows)}")
    print(f"- 最終資料筆數：{len(df)}")

    return df

def save_with_formatting(df):
    """儲存檔案並加上格式標示"""

    # 先存檔（不含格式）
    df.to_excel(OUTPUT_PATH, sheet_name=SHEET_NAME, index=False)

    # 用 openpyxl 加上樣式標示新增的資料列
    wb = load_workbook(OUTPUT_PATH)
    ws = wb[SHEET_NAME]

    # 找目標欄位的 1-based 欄號
    target_col_idx = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == TARGET_COL:
            target_col_idx = i
            break

    if target_col_idx:
        # 標示格式：淡綠底 + 藍字（表示分割處理過的資料）
        fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
        font = Font(color="0000FF")

        # 檢查每一列的序號欄，如果包含"-"則標示
        for row in range(2, ws.max_row + 1):
            seq_cell = ws.cell(row=row, column=1)  # A欄：序號
            if seq_cell.value and "-" in str(seq_cell.value):
                # 標示整列
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = fill
                    cell.font = font

    wb.save(OUTPUT_PATH)

def main():
    """主程式"""
    try:
        print("開始處理分割作業...")
        print(f"輸入檔：{INPUT_PATH}")

        # 處理分割資料
        df_result = process_split_data()

        if df_result is not None:
            # 儲存結果並加上格式
            save_with_formatting(df_result)

            print("\n分割作業完成！")
            print(f"輸出檔：{OUTPUT_PATH}")
        else:
            print("處理失敗，請檢查輸入檔案和欄位設定")

    except FileNotFoundError:
        print(f"錯誤：找不到輸入檔案 {INPUT_PATH}")
    except Exception as e:
        print(f"處理過程發生錯誤：{e}")

if __name__ == "__main__":
    main()